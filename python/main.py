import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(file_name):
    wb = xl.load_workbook(file_name)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, 
            min_row=2, 
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(file_name)





# CAR GAME

# help_inp = input("Press 'help' to get a list of car operations: ")
# print("Below are some commands")
# print("-"*len("Below are some commands"))
# print("start - to start the car")
# print("stop - to stop the car")
# print("quit - to exit")

# while help_inp.lower() == 'help':
#     cmd_now = input(">").lower().strip()
    
#     if cmd_now == "start":
#         print("Car started...Ready to go!")
#     elif cmd_now == "stop":
#         print("Car stopped")
#     elif cmd_now == "quit":
#         break
#     else:
#         print("I don't understand that...")




# GUESS GAME

# secret_number = 9
# guess_count = 0
# guess_limit = 3

# while guess_count < guess_limit:
#     guess = int(input("Guess: "))
#     guess_count += 1
    
#     if guess == secret_number:
#         print("You won!")
#         break

# else:
#     print("Sorry, You failed!")




# WEIGHT CONVERTER

# print("Weight Converter")
# print("-"*len("Weight Converter"))

# user_weight = int(input("Weight: "))

# user_unit = str(input("(L)bs or (K)g: "))

# if user_unit.lower() == "l":
#     print(f"{user_weight * 0.45}pounds")
# elif user_unit.lower() == "k":
#     print(f"{user_weight/0.45}kilos")
# else:
#     print("you didn't pick the right choice of unit!")